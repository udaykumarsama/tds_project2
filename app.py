from fastapi import FastAPI, Form, Query, HTTPException, File, UploadFile
from fastapi.middleware.cors import CORSMiddleware
import requests, os, tempfile, zipfile, csv, io, json, shutil, subprocess, sys, platform
from typing import Optional, Dict, Any, List
import numpy as np
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
from dotenv import load_dotenv
import PyPDF2
import pandas as pd
import re
import tabula
import markdown
import subprocess
import openpyxl
from openpyxl.utils import get_column_letter
import gc

# Load environment variables
load_dotenv()

app = FastAPI()

# Add CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# System prompts
TDS_SYSTEM_PROMPT = "Provide only the exact answer without any explanations, reasoning, or additional text. Be extremely concise."

CODE_GENERATION_PROMPT = "Generate only the exact code needed to solve the problem. No explanations, comments, or additional text. Return ONLY the executable code."

# API token
API_TOKEN = os.getenv("LLMFOUNDRY_TOKEN")
API_URL = "https://llmfoundry.straive.com/openai/v1/chat/completions"

def load_question_data():
    with open("data.json", "r", encoding="utf-8") as f:
        return json.load(f)

def find_similar_question(query, questions_data):
    questions = [item["question"] for item in questions_data]
    vectorizer = TfidfVectorizer()
    tfidf_matrix = vectorizer.fit_transform(questions)
    query_vector = vectorizer.transform([query])
    similarities = cosine_similarity(query_vector, tfidf_matrix).flatten()
    most_similar_idx = np.argmax(similarities)
    return questions_data[most_similar_idx], similarities[most_similar_idx]

def execute_code(code: str, code_type: str, working_dir: str, processed_files: List[str] = None) -> Dict[str, Any]:
    """Execute Python code or Git Bash command and return the result"""
    result = {"success": False, "output": "", "error": ""}
    
    # Add file handling code to ensure proper cleanup
    code_with_cleanup = code
    if code_type.lower() == "python" and "openpyxl" in code:
        # Add try-finally block to ensure Excel workbooks are closed
        if "workbook =" in code and "workbook.close()" not in code:
            code_with_cleanup = """
import gc
try:
""" + code.replace("\n", "\n    ") + """
finally:
    # Force garbage collection to release file handles
    gc.collect()
"""
    
    try:
        if code_type.lower() == "python":
            # Save and execute Python code
            script_path = os.path.join(working_dir, "temp_script.py")
            with open(script_path, "w", encoding="utf-8") as f:
                f.write(code_with_cleanup)
            
            process = subprocess.run(
                [sys.executable, script_path],
                cwd=working_dir,
                capture_output=True,
                text=True,
                timeout=30
            )
        else:
            # Execute Git Bash command
            shell = True
            if platform.system() == "Windows":
                git_bash_paths = [
                    r"C:\Program Files\Git\bin\bash.exe",
                    r"C:\Program Files (x86)\Git\bin\bash.exe"
                ]
                git_bash_exe = next((path for path in git_bash_paths if os.path.exists(path)), None)
                shell = [git_bash_exe, "-c"] if git_bash_exe else True
            else:
                shell = ["/bin/bash", "-c"]
            
            process = subprocess.run(
                code if shell is True else shell + [code],
                cwd=working_dir,
                shell=shell is True,
                capture_output=True,
                text=True,
                timeout=30
            )
        
        if process.returncode == 0:
            result["success"] = True
            result["output"] = process.stdout.strip()
        else:
            result["error"] = process.stderr.strip()
            
    except subprocess.TimeoutExpired:
        result["error"] = "Execution timed out after 30 seconds"
    except Exception as e:
        result["error"] = str(e)
        
    # Force garbage collection to help release file handles
    gc.collect()
        
    return result

def call_llm_api(system_prompt: str, user_prompt: str, model: str = "gpt-4o-mini") -> Dict[str, Any]:
    """Helper function to call the LLM API"""
    try:
        print(f"\n=== NEW API REQUEST ===")
        
        response = requests.post(
            API_URL,
            headers={"Authorization": f"Bearer {API_TOKEN}:tds-project2"},
            json={
                "model": model, 
                "messages": [
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": user_prompt}
                ]
            }
        )
        
        response_json = response.json()
        
        if "error" in response_json:
            print(f"\n=== API ERROR ===")
            print(f"{response_json['error'].get('message', 'Unknown API error')}")
            return {
                "success": False,
                "error": f"API Error: {response_json['error'].get('message', 'Unknown API error')}"
            }
        
        content = response_json["choices"][0]["message"]["content"].strip()
        print(f"\n=== API RESPONSE ===")
        print(f"{content}")
        
        return {
            "success": True,
            "content": content
        }
    except Exception as e:
        print(f"\n=== API ERROR ===")
        print(f"{str(e)}")
        return {
            "success": False,
            "error": f"API Request Error: {str(e)}"
        }

def predict_code_outcome(code: str, error_message: str, question: str, code_type: str) -> Dict[str, Any]:
    """When code execution fails, ask the LLM to predict the outcome"""
    prompt = f"""
Question that needs to be answered: "{question}"

The following {code_type} code was generated to answer this question:
```
{code}
```

The code execution failed with this error:
```
{error_message}
```

Based on the question and the code, what would the exact output have been if the code had executed successfully? Provide ONLY the raw output that would have been produced, with no explanations or additional text.
"""
    
    result = call_llm_api(
        "Predict only the exact output without any explanations or additional text.",
        prompt
    )
    
    if result["success"]:
        return {"success": True, "predicted_output": result["content"]}
    else:
        return {"success": False, "predicted_output": f"Failed to predict outcome: {result.get('error', 'Unknown error')}"}

def correct_code(code: str, error_message: str, question: str, code_type: str) -> Dict[str, Any]:
    """Ask the LLM to correct code that failed to execute"""
    prompt = f"""
Question that needs to be answered: "{question}"

The following {code_type} code was generated to answer this question:
```
{code}
```

The code execution failed with this error:
```
{error_message}
```

Please correct the code to fix the error and make it run successfully. Return ONLY the corrected code without any explanations or additional text.
"""
    
    result = call_llm_api(
        "Generate only the exact corrected code without any explanations or additional text.",
        prompt
    )
    
    if result["success"]:
        corrected_code = result["content"].strip()
        
        # Extract code from markdown code blocks if present
        if "```" in corrected_code:
            code_blocks = []
            lines = corrected_code.split("\n")
            in_code_block = False
            current_block = []
            
            for line in lines:
                if line.startswith("```"):
                    if in_code_block:
                        in_code_block = False
                        if current_block:
                            code_blocks.append("\n".join(current_block))
                            current_block = []
                    else:
                        in_code_block = True
                elif in_code_block:
                    current_block.append(line)
            
            if code_blocks:
                # Use the largest code block
                corrected_code = max(code_blocks, key=len)
        
        return {"success": True, "corrected_code": corrected_code}
    else:
        return {"success": False, "error": result.get("error", "Failed to correct code")}

def process_file_context(file_path, temp_dir, processed_files=None):
    """Process uploaded file and extract context"""
    if processed_files is None:
        processed_files = []
    
    file_context = ""
    filename = os.path.basename(file_path)
    
    # Process based on file type
    if filename.endswith('.zip'):
        # Extract ZIP and process contents
        with zipfile.ZipFile(file_path, 'r') as zip_ref:
            zip_ref.extractall(temp_dir)
        
        # Add extracted files to processed_files
        for root, _, files in os.walk(temp_dir):
            for f in files:
                if f != filename:
                    extracted_path = os.path.join(root, f)
                    processed_files.append(extracted_path)
                    
                    # Process CSV files
                    if f.endswith('.csv'):
                        file_context += process_csv_file(extracted_path, f)
                    # Process PDF files
                    elif f.endswith('.pdf'):
                        file_context += process_pdf_file(extracted_path, f)
                    # Process Excel files
                    elif f.endswith(('.xlsx', '.xls', '.xlsm')):
                        file_context += process_excel_file(extracted_path, f)
                    # Process JSONL files
                    elif f.endswith('.jsonl'):
                        file_context += process_jsonl_file(extracted_path, f)
    
    elif filename.endswith('.csv'):
        file_context += process_csv_file(file_path, filename)
    
    elif filename.endswith('.pdf'):
        file_context += process_pdf_file(file_path, filename)
    
    elif filename.endswith(('.xlsx', '.xls', '.xlsm')):
        file_context += process_excel_file(file_path, filename)
    
    elif filename.endswith('.jsonl'):
        file_context += process_jsonl_file(file_path, filename)
    
    else:
        # For other file types, read as text if possible
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read()
                file_context += f"\nFile: {filename}\nContent (first 1000 chars):\n{content[:1000]}\n"
                if len(content) > 1000:
                    file_context += "...(truncated)...\n"
        except Exception as e:
            file_context += f"\nFile: {filename} (could not read: {str(e)})\n"
    
    return file_context

def process_csv_file(file_path, filename):
    """Process a CSV file and return context"""
    context = ""
    try:
        with open(file_path, 'r', newline='', encoding='utf-8') as f:
            csv_reader = csv.reader(f)
            headers = next(csv_reader)
            
            context += f"\nFile: {filename}\nCSV columns: {', '.join(headers)}\n"
            
            # Read sample rows
            rows = []
            for i, row in enumerate(csv_reader):
                if i < 10:
                    rows.append(row)
                else:
                    break
            
            if rows:
                context += "Sample data:\n"
                for i, row in enumerate(rows):
                    context += f"Row {i+1}: {row}\n"
    except Exception as e:
        context += f"\nError reading CSV file {filename}: {str(e)}\n"
    
    return context

def process_pdf_file(file_path, filename):
    """Process a PDF file and return context"""
    context = f"\nFile: {filename} (PDF)\n"
    
    try:
        # Extract text content
        pdf_text = extract_pdf_text(file_path)
        context += f"Text Content (first 1000 chars):\n{pdf_text[:1000]}\n"
        if len(pdf_text) > 1000:
            context += "...(text content truncated)...\n"
        
        # Try to extract tables if they exist
        try:
            tables = extract_pdf_tables(file_path)
            if tables:
                context += f"\nDetected {len(tables)} tables in the PDF.\n"
                for i, table in enumerate(tables[:2]):  # Show only first 2 tables
                    context += f"\nTable {i+1} (preview):\n"
                    context += str(table.head(3).to_string()) + "\n"
                if len(tables) > 2:
                    context += f"\n...(additional {len(tables)-2} tables not shown)...\n"
        except Exception as e:
            context += f"\nTable extraction attempted but failed: {str(e)}\n"
    
    except Exception as e:
        context += f"Error processing PDF: {str(e)}\n"
    
    return context

def extract_pdf_text(file_path):
    """Extract text from a PDF file"""
    text = ""
    try:
        with open(file_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            for page_num in range(len(pdf_reader.pages)):
                page = pdf_reader.pages[page_num]
                text += page.extract_text() + "\n"
    except Exception as e:
        text = f"Error extracting text: {str(e)}"
    return text

def extract_pdf_tables(file_path):
    """Extract tables from a PDF file using tabula-py"""
    try:
        # Try to extract tables using tabula
        tables = tabula.read_pdf(file_path, pages='all', multiple_tables=True)
        return tables
    except Exception:
        # If tabula fails, try PyPDF2 and some basic table detection
        tables = []
        text = extract_pdf_text(file_path)
        
        # Very basic table detection - look for patterns that might indicate tables
        # This is a simplified approach and won't work for all tables
        table_patterns = re.findall(r'((?:[^\n]+\|[^\n]+\n){2,})', text)
        
        for pattern in table_patterns:
            lines = pattern.strip().split('\n')
            data = [line.split('|') for line in lines]
            if data and all(len(row) == len(data[0]) for row in data):
                df = pd.DataFrame(data[1:], columns=data[0])
                tables.append(df)
        
        return tables

def convert_pdf_to_markdown(file_path):
    """Convert PDF to Markdown format"""
    # Extract text from PDF
    text = extract_pdf_text(file_path)
    
    # Basic conversion of text to markdown
    # This is a simplified approach - real-world would need more sophisticated parsing
    md_text = text
    
    # Add headers (assuming lines with fewer than 50 chars that don't end with punctuation might be headers)
    lines = md_text.split('\n')
    for i, line in enumerate(lines):
        line = line.strip()
        if line and len(line) < 50 and not line[-1] in '.,:;?!)':
            # Check if it might be a header
            if i == 0 or not lines[i-1].strip():
                lines[i] = f'## {line}'
            
    # Detect lists
    for i, line in enumerate(lines):
        if re.match(r'^\s*[\d]+\.', line):
            lines[i] = re.sub(r'^\s*([\d]+\.)', r'\1', line)
        elif re.match(r'^\s*[•\-*]', line):
            lines[i] = re.sub(r'^\s*[•\-*]', '-', line)
    
    md_text = '\n'.join(lines)
    
    # Format with prettier if available
    try:
        with tempfile.NamedTemporaryFile(suffix='.md', delete=False) as temp_file:
            temp_file.write(md_text.encode('utf-8'))
            temp_path = temp_file.name
        
        # Try to format with prettier
        try:
            subprocess.run(['npx', 'prettier@3.4.2', '--write', temp_path], 
                          check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
            with open(temp_path, 'r', encoding='utf-8') as f:
                md_text = f.read()
        except Exception:
            # If prettier fails, continue with unformatted markdown
            pass
        
        # Clean up temp file
        if os.path.exists(temp_path):
            os.unlink(temp_path)
            
    except Exception as e:
        print(f"Error formatting markdown: {str(e)}")
    
    return md_text

def process_excel_file(file_path, filename):
    """Process an Excel file and return context"""
    context = f"\nFile: {filename} (Excel)\n"
    
    try:
        # Load the workbook with read_only mode to reduce memory usage
        workbook = None
        try:
            workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            
            # Get sheet names
            sheet_names = workbook.sheetnames
            context += f"Sheets: {', '.join(sheet_names)}\n\n"
            
            # Process each sheet (limit to first 3 sheets if there are many)
            for sheet_name in sheet_names[:3]:
                sheet = workbook[sheet_name]
                context += f"Sheet: {sheet_name}\n"
                
                # Get dimensions
                max_row = min(sheet.max_row, 10)  # Limit to first 10 rows
                max_col = min(sheet.max_column, 10)  # Limit to first 10 columns
                
                if max_row == 0 or max_col == 0:
                    context += "Sheet is empty\n\n"
                    continue
                    
                context += f"Dimensions: {sheet.max_row} rows x {sheet.max_column} columns\n"
                
                # Get column headers (first row)
                headers = []
                for col in range(1, max_col + 1):
                    cell_value = sheet.cell(row=1, column=col).value
                    headers.append(str(cell_value) if cell_value is not None else "")
                
                context += f"Headers: {', '.join(headers)}\n"
                
                # Sample data (first few rows)
                context += "Sample data:\n"
                
                # Create a small dataframe for display
                data = []
                for row in range(2, max_row + 1):  # Skip header row
                    row_data = []
                    for col in range(1, max_col + 1):
                        cell_value = sheet.cell(row=row, column=col).value
                        row_data.append(str(cell_value) if cell_value is not None else "")
                    data.append(row_data)
                
                if data:
                    df = pd.DataFrame(data, columns=headers)
                    context += df.to_string(index=False) + "\n\n"
                else:
                    context += "No data rows found\n\n"
            
            # If there are more sheets, mention them
            if len(sheet_names) > 3:
                context += f"...and {len(sheet_names) - 3} more sheets not shown\n"
        finally:
            # Ensure workbook is closed
            if workbook:
                workbook.close()
                
    except Exception as e:
        context += f"Error processing Excel file: {str(e)}\n"
    
    return context

def process_jsonl_file(file_path, filename):
    """Process a JSONL file and return context"""
    context = f"\nFile: {filename} (JSONL)\n"
    
    try:
        # Read the JSONL file
        records = []
        with open(file_path, 'r', encoding='utf-8') as f:
            for i, line in enumerate(f):
                if i < 10:  # Limit to first 10 records
                    try:
                        record = json.loads(line.strip())
                        records.append(record)
                    except json.JSONDecodeError:
                        context += f"Warning: Line {i+1} is not valid JSON\n"
                else:
                    break
        
        # Get total number of records
        with open(file_path, 'r', encoding='utf-8') as f:
            total_records = sum(1 for _ in f)
        
        context += f"Total records: {total_records}\n"
        
        if records:
            # Analyze structure of the first record to understand the data
            first_record = records[0]
            if isinstance(first_record, dict):
                keys = list(first_record.keys())
                context += f"Fields: {', '.join(keys)}\n\n"
            
            # Display sample records
            context += "Sample records:\n"
            for i, record in enumerate(records):
                context += f"Record {i+1}: {json.dumps(record, indent=2)}\n"
            
            if total_records > 10:
                context += f"\n...and {total_records - 10} more records not shown\n"
        else:
            context += "No valid records found in the file\n"
    
    except Exception as e:
        context += f"Error processing JSONL file: {str(e)}\n"
    
    return context

def generate_code_for_question(question: str, code_type: str, file_context: str, temp_dir: str, processed_files: List[str] = None) -> Dict[str, Any]:
    """Generate and execute code for a question"""
    print(f"\n=== GENERATING CODE ===")
    print(f"Code type: {code_type}")
    
    prompt = f"""
Question that needs to be answered: "{question}"

Context from uploaded files:
{file_context}

Task: Generate {code_type} code that will solve this question and produce the exact answer.
Requirements:
1. The code must be complete and executable
2. Use standard libraries when possible which doesn't require installation
3. Handle any file operations correctly
4. Return only the final answer as output
5. Do not include any explanatory text or comments in the output

Operating System: {platform.system()}
Python Version: {platform.python_version()}
Current working directory: {temp_dir}
"""
    
    result = call_llm_api(CODE_GENERATION_PROMPT, prompt)
    
    if not result["success"]:
        return {"success": False, "error": result["error"]}
    
    generated_code = result["content"].strip()
    
    # Extract code from markdown code blocks if present
    if "```" in generated_code:
        code_blocks = []
        lines = generated_code.split("\n")
        in_code_block = False
        current_block = []
        
        for line in lines:
            if line.startswith("```"):
                if in_code_block:
                    in_code_block = False
                    if current_block:
                        code_blocks.append("\n".join(current_block))
                        current_block = []
                else:
                    in_code_block = True
            elif in_code_block:
                current_block.append(line)
        
        if code_blocks:
            # Use the largest code block
            generated_code = max(code_blocks, key=len)
    
    # Execute the generated code with retry logic
    if processed_files is None:
        processed_files = []
    
    current_code = generated_code
    max_retries = 3
    retry_count = 0
    
    while retry_count < max_retries:
        execution_result = execute_code(current_code, code_type, temp_dir, processed_files)
        
        if execution_result["success"]:
            return {
                "success": True,
                "code": current_code,
                "result": execution_result["output"].strip()
            }
        else:
            retry_count += 1
            print(f"\n=== CODE EXECUTION FAILED (Attempt {retry_count}/{max_retries}) ===")
            print(f"Error: {execution_result['error']}")
            
            if retry_count < max_retries:
                # Try to correct the code
                print(f"\n=== ATTEMPTING TO CORRECT CODE ===")
                correction_result = correct_code(current_code, execution_result["error"], question, code_type)
                
                if correction_result["success"]:
                    current_code = correction_result["corrected_code"]
                    print(f"\n=== CODE CORRECTED, RETRYING EXECUTION ===")
                else:
                    print(f"\n=== FAILED TO CORRECT CODE: {correction_result.get('error', 'Unknown error')} ===")
                    break  # Exit the retry loop if we can't correct the code
            else:
                print(f"\n=== MAX RETRIES REACHED ===")
    
    # If we've exhausted all retries or correction failed, try to predict the outcome
    print(f"\n=== PREDICTING OUTCOME ===")
    prediction_result = predict_code_outcome(current_code, execution_result["error"], question, code_type)
    
    if "predicted_output" in prediction_result and prediction_result["predicted_output"]:
        print("\n=== USING PREDICTED OUTCOME ===")
        return {
            "success": True,
            "code": current_code,
            "result": prediction_result["predicted_output"],
            "note": "Execution failed after multiple attempts, this is a predicted result"
        }
    else:
        return {
            "success": False,
            "code": current_code,
            "error": execution_result["error"],
            "predicted_output": prediction_result.get("predicted_output", "Could not predict outcome")
        }

@app.post("/api")
async def answer_question_post(
    question: str = Form(...),
    file: Optional[UploadFile] = File(None)
):
    """Main API endpoint to process questions and files"""
    print("\n=== NEW API REQUEST ===")
    print(f"Question: {question}")
    
    # Create temporary directory for file processing
    temp_dir = tempfile.mkdtemp()
    try:
        file_context = ""
        processed_files = []
        
        # Process uploaded file if present
        if file and file.filename:
            print(f"Processing file: {file.filename}")
            file_path = os.path.join(temp_dir, file.filename)
            
            # Save the uploaded file
            try:
                with open(file_path, "wb") as buffer:
                    shutil.copyfileobj(file.file, buffer)
                
                processed_files.append(file_path)
                file_context = process_file_context(file_path, temp_dir, processed_files)
            except Exception as e:
                print(f"Error processing uploaded file: {str(e)}")
                return {"error": f"Error processing uploaded file: {str(e)}"}
        
        try:
            # Find most similar question in data.json
            questions_data = load_question_data()
            similar_question, similarity_score = find_similar_question(question, questions_data)
            
            # Process based on the matched question
            if similar_question['code'] == "yes":
                print("\n=== GENERATING CODE ===")
                # Generate and execute code
                code_type = similar_question['type'] or "python"
                print(f"Code type: {code_type}")
                result = generate_code_for_question(question, code_type, file_context, temp_dir, processed_files)
                
                if result["success"]:
                    print("\n=== CODE EXECUTION SUCCESS ===")
                    return {
                        "answer": result["result"],
                    }
                else:
                    # Check if we have a predicted outcome
                    if "predicted_output" in result and result["predicted_output"]:
                        print("\n=== USING PREDICTED OUTCOME ===")
                        return {
                            "answer": result["predicted_output"],
                            "note": "Execution failed, this is a predicted result"
                        }
                    else:
                        print("\n=== CODE EXECUTION FAILED ===")
                        print(f"Error: {result['error']}")
                        return {
                            "error": result["error"],
                            "code": result.get("code", ""),
                        }
            else:
                print("\n=== GETTING DIRECT ANSWER ===")
                # Get direct answer from LLM
                detailed_prompt = f"""
Question that needs to be answered: "{question}"

Context from uploaded files:
{file_context}

Please provide a clear and concise answer to the question.
"""
                answer_result = call_llm_api(TDS_SYSTEM_PROMPT, detailed_prompt)
                
                if answer_result["success"]:
                    return {"answer": answer_result["content"]}
                else:
                    return {"error": answer_result["error"]}
                
        except Exception as e:
            print(f"Error processing question: {str(e)}")
            return {"error": f"Error processing question: {str(e)}"}
    finally:
        # Clean up temporary files
        try:
            # Force close any open file handles
            gc.collect()
            
            # Try to remove the temporary directory
            shutil.rmtree(temp_dir, ignore_errors=True)
        except Exception as e:
            print(f"Warning: Could not clean up temporary directory: {str(e)}")

@app.get("/")
def root():
    return {"message": "TDS Project API is running. Use /api endpoint with POST requests."}